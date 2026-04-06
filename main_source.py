import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io

st.set_page_config(page_title="Logistics Quote Generator", layout="wide")

st.title("📦 Smart Logistics Quote Pipeline")
st.markdown("This version automatically finds the data, even if the number of pallets changes.")

# --- SIDEBAR: MANUAL INPUTS ---
with st.sidebar:
    st.header("Shipment Details")
    destination = st.text_input("Destination Address/City", "UK - Radial FAO Monat...")
    service = st.selectbox("Service", ["LCL", "LTL", "FCL", "Air", "Courier"])
    incoterms = st.selectbox("Incoterms", ["DAP", "DDP", "EXW", "FOB", "CIF"])
    commodity = st.text_input("Commodity", "Finished goods / Haircare")
    cargo_value = st.text_input("Value of Cargo", "USD$ 30,000.00")

# --- MAIN: FILE UPLOAD ---
col1, col2 = st.columns(2)
with col1:
    packing_file = st.file_uploader("Upload Packing List (.xlsx)", type=['xlsx'])
with col2:
    template_file = st.file_uploader("Upload Your Quote Template (.xlsx)", type=['xlsx'])

if packing_file and template_file:
    # --- DYNAMIC DATA EXTRACTION ---
    # We read the excel and skip the top branding rows to find the header
    df = pd.read_excel(packing_file, header=2) 
    
    # We clean the data to remove any 'Total' rows the outbound team might have added
    # We only want rows that actually have a PO Number
    df_clean = df.dropna(subset=['P.O.'])
    
    # CALCULATIONS (This works whether you have 1 pallet or 100)
    # Using the exact headers from your file: 'Tot. Weight / Bxs' and 'CBM'
    # Note: If CBM isn't a column, we can calculate it from dimensions!
    total_weight_kg = df_clean['Tot. Weight / Bxs'].sum()
    total_pallets = df_clean['PALLET QTY'].nunique() # Counts unique pallet IDs
    po_numbers = ", ".join(df_clean['P.O.'].unique().astype(str))

    st.success(f"✅ Extracted data for {total_pallets} Pallets across POs: {po_numbers}")

    # --- EXCEL FILLING ---
    template_bytes = template_file.read()
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active 

    # MAPPING TO YOUR TEMPLATE (Adjust cell letters as needed)
    ws['B2'] = destination
    ws['B3'] = service
    ws['B5'] = total_pallets
    ws['B11'] = f"{total_weight_kg:.2f} KGS"
    ws['B12'] = commodity
    ws['B13'] = incoterms
    ws['B14'] = cargo_value

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    
    # --- EMAIL DRAFT ---
    email_body = f"""
Subject: Quote Request - {service} - {destination}

Hi Team,

Please provide a quote for the following shipment:
- POs: {po_numbers}
- Destination: {destination}
- Service: {service}
- Total Pallets: {total_pallets}
- Total Weight: {total_weight_kg:.2f} KGS
- Incoterm: {incoterms}

Packing list is attached. Thanks!
    """

    # --- UI RESULTS ---
    st.divider()
    res_left, res_right = st.columns(2)
    
    with res_left:
        st.subheader("1. Download Your Form")
        st.download_button(
            label="📥 Download Quote Request",
            data=output.getvalue(),
            file_name=f"Quote_Request_{po_numbers[:10]}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    with res_right:
        st.subheader("2. Copy Email")
        st.text_area("Email Script:", value=email_body, height=200)

else:
    st.warning("Please upload both files to generate the quote.")
